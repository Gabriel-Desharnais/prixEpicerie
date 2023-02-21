#!/bin/python3
# -*- coding: utf-8 -*-

import pandas
import maxi
from openpyxl import Workbook
from openpyxl import styles

wb = Workbook()
# Activer le classeur
wb.active
# Créer la feuille du rapport
rapport = wb.create_sheet("Prix",0)


products = pandas.read_csv("productTable.csv")
maxiProducts = pandas.read_csv("maxi/productMaxi.csv")

productToExport = products.loc[products["maxi"] == True, ["uuid", "nom", "sku", "marque"]]
rapport.cell(1,1, "nom")
rapport.cell(1,2, "marque")
rapport.cell(1,3, "sku")
rapport.cell(1,4, "prix")
rapport.cell(1,5, "unité")
for i, prod in enumerate(productToExport.values):
	u = prod[0]
	m = maxiProducts.loc[maxiProducts["uuid"] == u]
	rapport.cell(i+2,1, prod[1])
	rapport.cell(i+2,2, prod[3])
	rapport.cell(i+2,3, prod[2])
	rapport.cell(i+2,4, m["prix"].values[0])
	rapport.cell(i+2,5, m["unité"].values[0])

wb.save("result/prix2023_02_21.xlsx")