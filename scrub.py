#!/bin/python3
# -*- coding: utf-8 -*-
 
"""Ce script utilise le module maxi pour importer les données scrubé sur le site de maxi"""
from openpyxl import Workbook
from openpyxl import styles
import pandas

produitsEs = pandas.read_csv("produitses.csv",na_filter=False)
productTable = pandas.read_csv("productTable.csv",na_filter=False)
productEquivalency = pandas.read_csv("productEquivalency.csv")
productMaxi = pandas.read_csv("maxi/productMaxi.csv")


wb = Workbook()
# Activer le classeur
wb.active
# Créer la feuille du rapport
rapport = wb.create_sheet("Analyse",0)
# Ajouter l'entête du tableau
rapport["A1"] = "Produits ?picerie solidaire"
rapport["J1"] = "Produits concurence"
rapport.merge_cells("A1:I1")
rapport.merge_cells("J1:U1")
rapport["A1"].fill = styles.PatternFill(start_color="bf819e", end_color="bf819e", fill_type="solid")
rapport["J1"].fill = styles.PatternFill(start_color="ffe994", end_color="ffe994", fill_type="solid")

rapport["A2"] = "Nom du produit"
rapport["B2"] = "Marque"
rapport["C2"] = "Numéro de produit interne"
rapport["D2"] = "Format"
rapport["E2"] = "Description"
rapport["F2"] = "Prix"
rapport["G2"] = "Unité"
rapport["H2"] = "Prix unitaire"
rapport["I2"] = "Unité"
rapport["J2"] = "Épicerie"
rapport["K2"] = "Nom du produit"
rapport["L2"] = "Marque"
rapport["M2"] = "Numéro de produit"
rapport["N2"] = "Format"
rapport["O2"] = "Description"
rapport["P2"] = "type de prix"
rapport["Q2"] = "Prix"
rapport["R2"] = "Unité"
rapport["S2"] = "Prix unitaire"
rapport["T2"] = "Unité"
rapport["U2"] = "Commentaire"

# Préparer la liste des produit à vérifier
pToLookUp = set(range(6))
#for prod in products_info.produitES:
#	pToLookUp |= products_info.equivalencyTable.get(prod, set())

#pInfo = maxi.scrub(pToLookUp)
#----------------- Simuler les réponses des épiceries ------------------
pInfo = []
for i, p in enumerate(pToLookUp):
	pInfo.append({"Épicerie": "Maxi",
				"nom": "lait 2%",
				"Marque": f"blabla{i}",
				"ndp":"sfsfweerwerewrwerfsfsdffffssfsdfewrwrresdfasdfsdafadf"[6*i:6*i+6],
				"Format":"1L",
				"description":"lait",
				"prix": str(2.15+0.10*i),
				"unité p": "ch",
				"prix unitaire": str((2.05+0.10*i)/10),
				"unité pu" :"/100ml"})

for i, p in enumerate(pToLookUp):
	pInfo.append({"Épicerie": "superC",
				"nom": "lait 2%",
				"Marque": f"blabla{i}",
				"ndp":"sfsfweerwerewrwerfsfsdffffssfsdfewrwrresdfasdfsdafadf"[6*i:6*i+6],
				"Format":"1L",
				"description":"lait",
				"prix": str(2.15+0.10*i),
				"unité p": "ch",
				"prix unitaire": str((2.15+0.10*i)/10),
				"unité pu" :"/100ml"})
#----------------- fin simulation ----------------

startRow = 3
for pes in produitsEs.values:
	uuid = pes[0]
	# Trouver le produit dans la bd producttable
	match = productTable.loc[productTable["uuid"] == uuid,["nom", "marque", "Format", "description"]].values
	if len(match) != 1:
		print("erreur nombre imprévu de match", len(match), uuid)
		continue
	match = match[0]
	# Aller chercher les équivalent
	equ = [] # Liste des produits équivanlents
	equNumber = 0
	# Ajouter le produit en tant que tel pour chaque épicerie
	nombreDepicerie = productTable.loc[productTable["uuid"] == uuid, ["maxi"]].values[0].tolist().count(True)
	if nombreDepicerie:
		equNumber += nombreDepicerie
		equ.append(uuid)
	equid = productEquivalency.loc[productEquivalency["uuid"] == uuid, ["equivalentUuid"]].values
	for u in equid:
		u = u[0]
		nombreDepicerie = productTable.loc[productTable["uuid"] == u, ["maxi"]].values[0].tolist().count(True)
		if nombreDepicerie:
			equNumber += nombreDepicerie
			equ.append(u)

	# Placer les informations sur le produit original
	endRow = startRow + max(len(equ), 1) - 1
	rapport.cell(startRow, 1, match[0]).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=1, end_row=endRow, end_column=1)
	rapport.cell(startRow, 2, match[1]).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=2, end_row=endRow, end_column=2)
	rapport.cell(startRow, 3, uuid).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=3, end_row=endRow, end_column=3)
	rapport.cell(startRow, 4, match[2]).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=4, end_row=endRow, end_column=4)
	rapport.cell(startRow, 5, match[3]).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=5, end_row=endRow, end_column=5)
	rapport.cell(startRow, 6, pes[1]).alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=6, end_row=endRow, end_column=6)
	rapport.cell(startRow, 7, "ch").alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=7, end_row=endRow, end_column=7)
	rapport.cell(startRow, 8, "").alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=8, end_row=endRow, end_column=8)
	rapport.cell(startRow, 9, "/100ml").alignment = styles.Alignment(vertical='center')
	rapport.merge_cells(start_row=startRow, start_column=9, end_row=endRow, end_column=9)

	try:
		esPrice = float(pes[1])
	except:
		esPrice = 100000000 # FIXME c'est pas comme ça qu'on fait ça
	# Ajouter les produits équivalents
	smallestPrice = None # should be none
	spl = []
	for i, u in enumerate(equ):
		u = u
		p = productTable.loc[productTable["uuid"]==u,["nom", "marque", "sku", "Format", "description"]].values[0]
		pm = productMaxi.loc[productMaxi["uuid"]==u, ["prix", "unité","type_de_prix","commentaire"]].values[0]
		fp = pInfo[i]
		rapport.cell(startRow+i, 10, "Maxi")
		rapport.cell(startRow+i, 11, p[0])
		rapport.cell(startRow+i, 12, p[1])
		rapport.cell(startRow+i, 13, p[2])
		rapport.cell(startRow+i, 14, p[3])
		rapport.cell(startRow+i, 15, p[4])
		rapport.cell(startRow+i, 16, pm[2])
		rapport.cell(startRow+i, 17, pm[0])
		rapport.cell(startRow+i, 18, pm[1])
		rapport.cell(startRow+i, 19, "-")
		rapport.cell(startRow+i, 20, "-")
		rapport.cell(startRow+i, 21, pm[3])
		# Vérifié si le moins chers
		try:
			price  = float(pm[0])
		except:
			price = 10000000 # FIXME c'est pas comme ça qu'on fait ça
		if smallestPrice is None or smallestPrice > price:
			#Nouveau meilleur prix
			smallestPrice = price
			spl = [startRow+i]
		elif smallestPrice == price:
			spl.append(startRow + i)




	side = styles.Side(border_style="thick", color="000000")
	border = styles.Border(top=side,left=side,right=side, bottom=side)

	sideToBorder = ["top", "right", "bottom", "left"]
	row = startRow
	column = 1
	for s in range(4):
		rapport.cell(row,column).border = styles.Border(**{sideToBorder[s]: side, sideToBorder[s-1]: side})
		if s%2:
			borderLen = max(len(equ), 1) - 2
		else:
			borderLen = 19
		match s:
			case 0:
				column += 1
			case 1:
				row += 1
			case 2:
				column -= 1
			case 3:
				row -= 1
		for i in range(borderLen):
			rapport.cell(row,column).border = styles.Border(**{sideToBorder[s]: side})
			match s:
				case 0:
					column += 1
				case 1:
					row += 1
				case 2:
					column -= 1
				case 3:
					row -= 1

	# Ajouter la couleur pour le produit le moins chers
	# Est-ce que es a le meilleur prix
	try:
		if esPrice <= smallestPrice:
			# Youpi on a le meilleur prix mettre le produit en vert
			couleur = styles.PatternFill(start_color="81d41a", end_color="81d41a", fill_type="solid")
		else:
			# on n'a pas le meilleur prix mettre le produit en rouge
			couleur = styles.PatternFill(start_color="ff3838", end_color="ff3838", fill_type="solid")
		# appliquer la couleur
		for i in range(1, 10):
			rapport.cell(startRow, i).fill = couleur
	except:
		pass

	# Appliquer la couleur aux meilleur prix équivalent
	couleur = styles.PatternFill(start_color="729fcf", end_color="729fcf", fill_type="solid")
	for row in spl:
		for i in range(10, 22):
			rapport.cell(row, i).fill = couleur
	
	# Augmenter la ligne de départ
	startRow = endRow + 1





wb.save("result/rapport.xlsx")